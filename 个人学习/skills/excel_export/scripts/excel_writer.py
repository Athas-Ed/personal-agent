"""写入 .xlsx（openpyxl）。"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _sheet_title(name: str) -> str:
    s = name.replace("\\", "-").replace("/", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
    return s[:31] if len(s) > 31 else s


def _header_row(ws, titles: List[str]) -> None:
    ws.append(titles)
    for c in ws[1]:
        c.font = Font(bold=True)


def write_dialogue_xlsx(
    path: Path,
    rows: Sequence[Tuple[str, str]],
    *,
    sheet_name: str = "对话",
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _sheet_title(sheet_name)
    _header_row(ws, ["序号", "角色", "台词"])
    for i, (spk, line) in enumerate(rows, start=1):
        ws.append([i, spk, line])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_task_xlsx(
    path: Path,
    rows: Sequence[Dict[str, str]],
    *,
    sheet_name: str = "任务",
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _sheet_title(sheet_name)
    _header_row(ws, ["序号", "任务名", "目标", "奖励", "摘要"])
    for i, r in enumerate(rows, start=1):
        ws.append(
            [
                i,
                r.get("name", ""),
                r.get("goal", ""),
                r.get("reward", ""),
                r.get("summary", ""),
            ]
        )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_table_xlsx(path: Path, header: List[str], data_rows: List[List[str]], *, sheet_name: str = "表格") -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _sheet_title(sheet_name)
    _header_row(ws, header)
    for row in data_rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_combined_xlsx(
    path: Path,
    *,
    dialogue_rows: Sequence[Tuple[str, str]] | None = None,
    task_rows: Sequence[Dict[str, str]] | None = None,
    tables: Sequence[Tuple[List[str], List[List[str]]]] | None = None,
) -> None:
    wb = Workbook()
    used_active = False

    def next_sheet(title: str):
        nonlocal used_active
        if not used_active:
            ws = wb.active
            used_active = True
        else:
            ws = wb.create_sheet()
        assert ws is not None
        ws.title = _sheet_title(title)
        return ws

    if dialogue_rows:
        ws = next_sheet("对话")
        _header_row(ws, ["序号", "角色", "台词"])
        for i, (spk, line) in enumerate(dialogue_rows, start=1):
            ws.append([i, spk, line])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if task_rows:
        ws = next_sheet("任务")
        _header_row(ws, ["序号", "任务名", "目标", "奖励", "摘要"])
        for i, r in enumerate(task_rows, start=1):
            ws.append(
                [
                    i,
                    r.get("name", ""),
                    r.get("goal", ""),
                    r.get("reward", ""),
                    r.get("summary", ""),
                ]
            )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if tables:
        for ti, (header, data_rows) in enumerate(tables, start=1):
            ws = next_sheet(f"表{ti}")
            _header_row(ws, header)
            for row in data_rows:
                ws.append(row)

    if not dialogue_rows and not task_rows and not tables:
        ws = wb.active
        assert ws is not None
        ws.append(["提示"])
        ws.append(["无导出数据"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
